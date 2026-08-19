"""
Genera la planilla editable CARGA_MANUAL_SEP.xlsx — mismo estilo que el flujo de caja
(celdas amarillas = las únicas que se editan, todo lo demás se calcula solo).

Tiene 4 hojas, todas con el mismo formato de columnas (RBD | Establecimiento | [Item] |
Enero..Diciembre | Total Año) para que se vean alineadas entre sí:

  - INGRESO         : Ingreso SEP por RBD, por mes. Sin ítem (una fila por RBD).
                      Semilla: el ingreso anual (Estado_RBD_2026.xlsx, columna TOTAL SEP)
                      repartido en 12 partes iguales — es un supuesto simple para partir;
                      edítalo si sabes el calendario real de transferencias MINEDUC.
  - REMUNERACIONES  : Gasto Subt.21 (remuneraciones) por RBD, por mes. Sin ítem.
                      Semilla: enero-julio = gasto real ya automatizado (CasChile,
                      FuenteCorregida); agosto-diciembre = el promedio mensual real,
                      como proyección editable (para que el total anual de partida sea
                      igual al que ya se mostraba en el panel).
  - SUBT22          : Gasto Subt.22 (bienes y servicios) por RBD, por ítem, por mes.
  - SUBT29          : Gasto Subt.29 (activos) por RBD, por ítem, por mes.

Uso: python3 build_carga_manual.py
Salida: carga_manual/CARGA_MANUAL_SEP.xlsx
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

d = json.load(open('consolidado_sep.json', encoding='utf-8'))
rbds = sorted([r for r in d['registros'] if r['marco_anual_proy'] > 0], key=lambda x: int(x['rbd']))

ingreso_sep = json.load(open('ingreso_sep_rbd.json', encoding='utf-8'))
gasto_rem = json.load(open('gasto_rem_sep.json', encoding='utf-8'))  # {periodo: {rbd: monto}}

ITEMS_22 = [
    'SALIDAS PEDAGOGICAS', 'MATERIAL PEDAGOGICO Y DIDACTICO', 'IMPRESORAS', 'VESTUARIO',
    'ALIMENTACIÓN', 'MEJORAMIENTO EN REDES', 'PRODUCCIONES Y/O RECONOCIMIENTOS, PREMIOS',
    'MATERIAL MUSICAL', 'MATERIAL DEPORTIVO', 'PLATAFORMA EDUCATIVA / LIBRO DIGITAL',
]
ITEMS_29 = ['MOBILIARIO', 'MAQUINAS Y EQUIPOS', 'EQUIPOS INFORMATICOS', 'PROGRAMAS INFORMATICOS']

MESES = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO',
         'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
PERIODOS_REALES = ['202601', '202602', '202603', '202604', '202605', '202606', '202607']
MES_DE_PERIODO = dict(zip(PERIODOS_REALES, MESES[:7]))

FONT_NAME = 'Arial'
HEADER_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
HEADER_FILL = PatternFill('solid', fgColor='2A78D6')
LABEL_FONT = Font(name=FONT_NAME, size=10)
LABEL_FONT_BOLD = Font(name=FONT_NAME, size=10, bold=True)
INPUT_FILL = PatternFill('solid', fgColor='FFFFCC')
INPUT_FONT = Font(name=FONT_NAME, size=10, color='0000FF')
TOTAL_FONT = Font(name=FONT_NAME, size=10, bold=True)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = '$#,##0;($#,##0);-'

# Columnas fijas e idénticas en las 4 hojas para que queden alineadas:
# RBD(A, 8) | Establecimiento(B, 34) | Item(C, 30) | Ene..Dic (D..O, 11 c/u) | Total Año(P, 13)
# Las hojas sin ítem (INGRESO, REMUNERACIONES) dejan la columna C vacía/oculta con el mismo ancho,
# para que las 4 tablas tengan exactamente el mismo grid visual.
COL_WIDTHS = {'A': 8, 'B': 34, 'C': 30}


def seed_ingreso(rbd):
    anual = ingreso_sep.get(rbd, 0.0)
    base = round(anual / 12)
    valores = [base] * 12
    # ajusta el último mes para que la suma cuadre exacto con el anual (evita arrastre por redondeo)
    valores[-1] = round(anual - base * 11)
    return valores


def seed_remuneraciones(rbd):
    reales = [gasto_rem.get(p, {}).get(rbd, 0.0) for p in PERIODOS_REALES]
    n = len(reales)
    promedio = round(sum(reales) / n) if n else 0
    valores = [round(v) for v in reales] + [promedio] * (12 - n)
    return valores


def build_sheet(wb, sheet_name, con_item, items, seed_fn):
    ws = wb.create_sheet(sheet_name)
    headers = ['RBD', 'Establecimiento', 'Item'] + MESES + ['Total Año']
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = 'D2'
    ws.row_dimensions[1].height = 30

    first_month_col = 4
    last_month_col_letter = get_column_letter(first_month_col + len(MESES) - 1)
    total_col = first_month_col + len(MESES)

    row = 2
    filas_por_rbd = items if con_item else [None]
    for r in rbds:
        seed = seed_fn(r['rbd'])
        for item in filas_por_rbd:
            ws.cell(row=row, column=1, value=int(r['rbd'])).font = LABEL_FONT
            ws.cell(row=row, column=2, value=r['nombre']).font = LABEL_FONT
            item_cell = ws.cell(row=row, column=3, value=item if item else '')
            item_cell.font = LABEL_FONT
            for mi, mes in enumerate(MESES):
                c = first_month_col + mi
                valor = seed[mi] if not con_item else 0
                cell = ws.cell(row=row, column=c, value=valor)
                cell.font = INPUT_FONT
                cell.fill = INPUT_FILL
                cell.number_format = MONEY_FMT
                cell.border = BORDER
            total_cell = ws.cell(row=row, column=total_col,
                                  value=f"=SUM(D{row}:{last_month_col_letter}{row})")
            total_cell.font = TOTAL_FONT
            total_cell.number_format = MONEY_FMT
            total_cell.border = BORDER
            for c in (1, 2, 3):
                ws.cell(row=row, column=c).border = BORDER
            row += 1

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    for mi in range(len(MESES)):
        ws.column_dimensions[get_column_letter(4 + mi)].width = 11
    ws.column_dimensions[get_column_letter(4 + len(MESES))].width = 13

    last_data_row = row - 1
    total_row = row + 1
    ws.cell(row=total_row, column=2, value='TOTAL').font = LABEL_FONT_BOLD
    for mi in range(len(MESES)):
        c = 4 + mi
        col_letter = get_column_letter(c)
        cell = ws.cell(row=total_row, column=c, value=f"=SUM({col_letter}2:{col_letter}{last_data_row})")
        cell.font = TOTAL_FONT
        cell.number_format = MONEY_FMT
    total_col_letter = get_column_letter(4 + len(MESES))
    cell = ws.cell(row=total_row, column=4 + len(MESES), value=f"=SUM({total_col_letter}2:{total_col_letter}{last_data_row})")
    cell.font = TOTAL_FONT
    cell.number_format = MONEY_FMT

    return ws


wb = openpyxl.Workbook()
wb.remove(wb.active)

ws_i = wb.create_sheet('INSTRUCCIONES')
ws_i.column_dimensions['A'].width = 100
instrucciones = [
    ('CARGA MANUAL SEP — Ingreso, Remuneraciones (Subt.21), Bienes y Servicios (Subt.22 y 29)', LABEL_FONT_BOLD),
    ('', LABEL_FONT),
    ('Hojas de esta planilla:', LABEL_FONT_BOLD),
    ('  • INGRESO: ingreso SEP por RBD, por mes.', LABEL_FONT),
    ('  • REMUNERACIONES: gasto en Subtítulo 21 (remuneraciones) por RBD, por mes.', LABEL_FONT),
    ('  • SUBT22 y SUBT29: gasto en bienes y servicios / activos, por RBD, por ítem y por mes.', LABEL_FONT),
    ('', LABEL_FONT),
    ('Cómo usar esta planilla:', LABEL_FONT_BOLD),
    ('1. Las celdas amarillas (Enero a Diciembre) son las únicas que debes editar. Todo lo demás (Total Año, fila TOTAL) se calcula solo — no lo edites.', LABEL_FONT),
    ('2. INGRESO parte con el ingreso anual repartido en 12 partes iguales (supuesto simple). Si conoces el calendario real de transferencias MINEDUC, ajusta cada mes.', LABEL_FONT),
    ('3. REMUNERACIONES parte con los meses de enero a julio iguales al gasto real ya cargado por el sistema (CasChile), y agosto a diciembre con el promedio mensual como proyección editable — puedes corregir cualquier mes cuando tengas el dato real.', LABEL_FONT),
    ('4. SUBT22 y SUBT29 parten en $0 — anótalos a medida que ejecutes el gasto cada mes.', LABEL_FONT),
    ('5. No agregues ni borres columnas de mes. Si necesitas un RBD o ítem nuevo, agrega una fila copiando el formato de una fila existente.', LABEL_FONT),
    ('6. Guarda el archivo con el mismo nombre en esta misma carpeta — el panel se vuelve a generar leyendo esta planilla (o puedes subir los cambios directo desde el dashboard, con el botón "Descargar carga").', LABEL_FONT),
]
for i, (text, font) in enumerate(instrucciones, start=1):
    cell = ws_i.cell(row=i, column=1, value=text)
    cell.font = font
    cell.alignment = Alignment(wrap_text=True, vertical='top')

build_sheet(wb, 'INGRESO', con_item=False, items=None, seed_fn=seed_ingreso)
build_sheet(wb, 'REMUNERACIONES', con_item=False, items=None, seed_fn=seed_remuneraciones)
build_sheet(wb, 'SUBT22', con_item=True, items=ITEMS_22, seed_fn=lambda rbd: [0] * 12)
build_sheet(wb, 'SUBT29', con_item=True, items=ITEMS_29, seed_fn=lambda rbd: [0] * 12)

out_path = 'carga_manual/CARGA_MANUAL_SEP.xlsx'
wb.save(out_path)
print(f"Guardado: {out_path}")
print(f"RBDs: {len(rbds)}")
print(f"Filas INGRESO/REMUNERACIONES: {len(rbds)} c/u; SUBT22: {len(rbds)*len(ITEMS_22)}; SUBT29: {len(rbds)*len(ITEMS_29)}")
