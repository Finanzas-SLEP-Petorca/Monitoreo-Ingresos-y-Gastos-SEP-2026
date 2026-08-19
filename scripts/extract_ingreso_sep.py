"""
Extrae el Ingreso SEP anual por RBD directamente desde el archivo oficial
CARGA/Estado_RBD_2026.xlsx (hoja 'ESTADO CTA E.E.'), columna 'TOTAL SEP'
(SEP Prioritarios + SEP Preferentes).

A diferencia del marco anterior, este valor:
  - Es un monto YA ANUAL (Ley Vigente 2026 tal como la tiene tu equipo), no se extrapola.
  - NO incluye Carrera Docente / CPEIP — según validaste, la regla 70/30 se cumple mejor
    solo con SEP (el gasto real en remuneraciones SEP queda ~73% de este ingreso, muy
    cerca del 70% esperado; incluyendo CPEIP completo bajaba a ~33%, y por eso se sacó).

Si el archivo Estado_RBD_2026.xlsx se actualiza (nueva proyección, nuevos RBD, etc.),
basta con volver a correr este script y luego consolidar.py.

Salida: ingreso_sep_rbd.json  { rbd: monto_anual }
"""
import json
import openpyxl

SRC = "/mnt/user-data/uploads/slep-petorca-finance/CARGA/Estado_RBD_2026.xlsx"
COL_RBD = 2
COL_NOMBRE = 3
COL_TOTAL_SEP = 36  # 'TOTAL SEP' = SEP_PRIORITARIOS + SEP_PREFERENTES


def to_f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb['ESTADO CTA E.E.']

    ingreso_sep = {}
    nombres = {}
    row = 13  # primera fila de datos real (fila 12 es encabezado)
    while True:
        rbd_val = ws.cell(row=row, column=COL_RBD).value
        if rbd_val is None or str(rbd_val).strip() == '':
            break
        rbd = str(int(rbd_val)) if isinstance(rbd_val, (int, float)) else str(rbd_val).strip()
        nombre = ws.cell(row=row, column=COL_NOMBRE).value
        total_sep = to_f(ws.cell(row=row, column=COL_TOTAL_SEP).value)
        ingreso_sep[rbd] = total_sep
        nombres[rbd] = nombre
        row += 1

    print(f"RBDs leídos: {len(ingreso_sep)} (filas 13 a {row - 1})")
    total = sum(ingreso_sep.values())
    print(f"TOTAL SEP anual (suma todos los RBD): ${total:,.0f}")

    json.dump(ingreso_sep, open('ingreso_sep_rbd.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
    print("Guardado: ingreso_sep_rbd.json")
    print(f"Fuente: {SRC}")


if __name__ == '__main__':
    main()
