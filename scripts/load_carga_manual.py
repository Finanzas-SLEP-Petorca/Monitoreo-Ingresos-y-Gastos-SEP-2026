"""
Loader de la planilla CARGA_MANUAL_SEP.xlsx (patrón flujo de caja overrides).

Lee las 4 hojas (INGRESO, REMUNERACIONES, SUBT22, SUBT29) y produce carga_manual_data.json
con:
  - ingreso_por_rbd:  {rbd: {nombre, meses:{MES:monto}, total}}
  - rem_por_rbd:      {rbd: {nombre, meses:{MES:monto}, total}}
  - agregado_por_rbd: {rbd: {nombre, gasto_subt22, gasto_subt29, gasto_bys, items:[...]}}
      (igual estructura que antes, para Subt.22 y Subt.29 con detalle por ítem/mes)

Uso: python3 load_carga_manual.py [ruta_al_xlsx]
Default: carga_manual/CARGA_MANUAL_SEP.xlsx
"""
import json
import sys
from datetime import datetime

import openpyxl

MESES = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO',
         'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']


def leer_hoja_sin_item(ws):
    """Lee INGRESO o REMUNERACIONES: una fila por RBD, sin dimensión de ítem."""
    out = {}
    row = 2
    while True:
        rbd_val = ws.cell(row=row, column=1).value
        if rbd_val is None or str(rbd_val).strip() == '':
            break
        rbd = str(int(rbd_val)) if isinstance(rbd_val, (int, float)) else str(rbd_val).strip()
        nombre = ws.cell(row=row, column=2).value
        meses = {}
        total = 0.0
        for mi, mes in enumerate(MESES):
            v = ws.cell(row=row, column=4 + mi).value
            v = float(v) if isinstance(v, (int, float)) else 0.0
            meses[mes] = v
            total += v
        out[rbd] = {'rbd': rbd, 'nombre': nombre, 'meses': meses, 'total': round(total)}
        row += 1
    return out


def leer_hoja_con_item(ws, subtitulo):
    """Lee SUBT22 o SUBT29: RBD + ítem por fila. Devuelve lista de registros de detalle
    (solo con monto != 0) para trazabilidad, más un dict de agregados intermedios."""
    detalle = []
    items_por_rbd = {}
    row = 2
    while True:
        rbd_val = ws.cell(row=row, column=1).value
        if rbd_val is None or str(rbd_val).strip() == '':
            break
        rbd = str(int(rbd_val)) if isinstance(rbd_val, (int, float)) else str(rbd_val).strip()
        nombre = ws.cell(row=row, column=2).value
        item = ws.cell(row=row, column=3).value
        meses = {}
        total = 0.0
        for mi, mes in enumerate(MESES):
            v = ws.cell(row=row, column=4 + mi).value
            v = float(v) if isinstance(v, (int, float)) else 0.0
            meses[mes] = v
            total += v
            if v != 0:
                detalle.append({'rbd': rbd, 'nombre': nombre, 'subtitulo': subtitulo,
                                 'item': item, 'mes': mes, 'mes_idx': mi + 1, 'monto': v})
        items_por_rbd.setdefault(rbd, {'rbd': rbd, 'nombre': nombre, 'items': []})
        items_por_rbd[rbd]['items'].append({'subtitulo': subtitulo, 'item': item,
                                             'meses': meses, 'total': round(total)})
        row += 1
    return detalle, items_por_rbd


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else 'carga_manual/CARGA_MANUAL_SEP.xlsx'
    wb = openpyxl.load_workbook(path, data_only=True)

    ingreso_por_rbd = leer_hoja_sin_item(wb['INGRESO']) if 'INGRESO' in wb.sheetnames else {}
    rem_por_rbd = leer_hoja_sin_item(wb['REMUNERACIONES']) if 'REMUNERACIONES' in wb.sheetnames else {}

    detalle_22, items_22 = leer_hoja_con_item(wb['SUBT22'], '22') if 'SUBT22' in wb.sheetnames else ([], {})
    detalle_29, items_29 = leer_hoja_con_item(wb['SUBT29'], '29') if 'SUBT29' in wb.sheetnames else ([], {})
    detalle = detalle_22 + detalle_29

    agregados = {}
    for rbd, info in items_22.items():
        ag = agregados.setdefault(rbd, {'rbd': rbd, 'nombre': info['nombre'],
                                         'gasto_subt22': 0.0, 'gasto_subt29': 0.0,
                                         'gasto_bys': 0.0, 'items': []})
        ag['items'].extend(info['items'])
    for rbd, info in items_29.items():
        ag = agregados.setdefault(rbd, {'rbd': rbd, 'nombre': info['nombre'],
                                         'gasto_subt22': 0.0, 'gasto_subt29': 0.0,
                                         'gasto_bys': 0.0, 'items': []})
        ag['items'].extend(info['items'])
    for ag in agregados.values():
        ag['gasto_subt22'] = sum(i['total'] for i in ag['items'] if i['subtitulo'] == '22')
        ag['gasto_subt29'] = sum(i['total'] for i in ag['items'] if i['subtitulo'] == '29')
        ag['gasto_bys'] = ag['gasto_subt22'] + ag['gasto_subt29']

    out = {
        'fuente_archivo': path,
        'fecha_carga': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'n_registros_detalle': len(detalle),
        'n_rbd_con_gasto': sum(1 for a in agregados.values() if a['gasto_bys'] > 0),
        'detalle': detalle,
        'agregado_por_rbd': agregados,
        'ingreso_por_rbd': ingreso_por_rbd,
        'rem_por_rbd': rem_por_rbd,
    }

    with open('carga_manual_data.json', 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"Leído: {path}")
    print(f"RBDs con ingreso cargado: {len(ingreso_por_rbd)} — total: ${sum(v['total'] for v in ingreso_por_rbd.values()):,.0f}")
    print(f"RBDs con remuneraciones cargadas: {len(rem_por_rbd)} — total: ${sum(v['total'] for v in rem_por_rbd.values()):,.0f}")
    print(f"RBDs con gasto Subt.22/29 cargado: {out['n_rbd_con_gasto']}")
    total_22 = sum(a['gasto_subt22'] for a in agregados.values())
    total_29 = sum(a['gasto_subt29'] for a in agregados.values())
    print(f"Total Subt22 cargado: ${total_22:,.0f}")
    print(f"Total Subt29 cargado: ${total_29:,.0f}")
    print("Guardado: carga_manual_data.json")


if __name__ == '__main__':
    main()
