import glob, re, json
from collections import defaultdict
import openpyxl

BASE = "/mnt/user-data/uploads/slep-petorca-finance/CARGA/REMUNERACIONES/Educacion P02"

def norm(s):
    s = s.upper().strip()
    s = re.sub(r'[ÁÀÄ]', 'A', s); s = re.sub(r'[ÉÈË]', 'E', s)
    s = re.sub(r'[ÍÌÏ]', 'I', s); s = re.sub(r'[ÓÒÖ]', 'O', s)
    s = re.sub(r'[ÚÙÜ]', 'U', s); s = re.sub(r'Ñ', 'N', s)
    s = re.sub(r'[^A-Z0-9 ]', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

nombre_to_rbd = json.load(open('nombre_to_rbd.json', encoding='utf-8'))

# gasto_sep[periodo][centro_costo_norm] = monto  (before mapping to RBD)
gasto_sep_raw = defaultdict(lambda: defaultdict(float))
unmatched_names = defaultdict(float)

files = sorted(glob.glob(f"{BASE}/*/*.xlsx"))
print(f"Total archivos centralizacion: {len(files)}")

for f in files:
    m = re.search(r'/(20\d{4})/', f)
    periodo = m.group(1) if m else None
    if periodo is None:
        print("  SKIP (sin periodo):", f)
        continue
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
    except Exception as e:
        print(f"  ERROR abriendo {f}: {e}")
        continue
    if 'FuenteCorregida' not in wb.sheetnames:
        print(f"  SKIP (sin hoja FuenteCorregida): {f}")
        continue
    ws = wb['FuenteCorregida']
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
    try:
        col_centro = headers.index('Centro_Costo') + 1
        col_fuente = headers.index('Fuente') + 1
        col_monto = headers.index('Monto_Corregido') + 1
    except ValueError:
        print(f"  SKIP (columnas no encontradas): {f} headers={headers[:10]}")
        continue
    n_sep_rows = 0
    for r in range(2, ws.max_row + 1):
        fuente = ws.cell(row=r, column=col_fuente).value
        if fuente != 'PROYECTO SEP':
            continue
        centro = ws.cell(row=r, column=col_centro).value
        monto = ws.cell(row=r, column=col_monto).value or 0
        if centro is None:
            continue
        gasto_sep_raw[periodo][norm(str(centro))] += float(monto)
        n_sep_rows += 1
    print(f"{f.split('/')[-1]}: periodo={periodo}, filas_SEP={n_sep_rows}")

# Map centro_costo -> rbd
gasto_sep_rbd = defaultdict(lambda: defaultdict(float))
for periodo, centros in gasto_sep_raw.items():
    for centro_norm, monto in centros.items():
        rbd = nombre_to_rbd.get(centro_norm)
        if rbd:
            gasto_sep_rbd[periodo][rbd] += monto
        else:
            unmatched_names[centro_norm] += monto

print("\nPeriodos gasto remuneraciones SEP:", sorted(gasto_sep_rbd.keys()))
for p in sorted(gasto_sep_rbd.keys()):
    total = sum(gasto_sep_rbd[p].values())
    print(f"  {p}: {len(gasto_sep_rbd[p])} RBDs, total = {total:,.0f}")

print(f"\nNombres NO mapeados a RBD ({len(unmatched_names)}), monto total sin mapear:", sum(unmatched_names.values()))
for name, monto in sorted(unmatched_names.items(), key=lambda x:-x[1])[:20]:
    print(f"  {name}: {monto:,.0f}")

json.dump({p: dict(v) for p, v in gasto_sep_rbd.items()}, open('gasto_rem_sep.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
json.dump(dict(unmatched_names), open('unmatched_names.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
